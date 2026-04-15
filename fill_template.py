import sys
import json
import traceback

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

def fill_template(template_path, output_path, data_path):
    try:
        with open(data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"ERROR reading JSON: {e}")
        sys.exit(1)

    items = data.get('items', [])
    wo_number = data.get('woNumber', '')

    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f"ERROR loading template: {e}")
        sys.exit(1)

    ws = wb.active

    start_row = 18
    for i, item in enumerate(items):
        row = start_row + i
        try:
            ws.cell(row=row, column=2, value=item.get('description', ''))
        except AttributeError:
            pass
        try:
            ws.cell(row=row, column=9, value=item.get('unit', ''))
        except AttributeError:
            pass
        try:
            qty = float(item.get('quantity', 1))
        except:
            qty = 1
        try:
            ws.cell(row=row, column=10, value=qty)
        except AttributeError:
            pass
        try:
            ws.cell(row=row, column=13, value=item.get('item_number', ''))
        except AttributeError:
            pass

    for r in range(32, 36):
        try:
            cell = ws.cell(row=r, column=7)
            cell.value = wo_number
        except AttributeError:
            # خلية مدموجة - تخطي
            pass

    try:
        wb.save(output_path)
        print('OK')
    except Exception as e:
        print(f"ERROR saving: {e}")
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(f"Usage: {sys.argv[0]} template output data.json")
        sys.exit(1)
    fill_template(sys.argv[1], sys.argv[2], sys.argv[3])
